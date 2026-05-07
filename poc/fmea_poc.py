#!/usr/bin/env python3
"""
Phase 0 PoC — FMEA 단일 생성 검증 스크립트
양식 기준: AIAG & VDA FMEA 1st Edition (2019) PFMEA 표준 컬럼 구조

실행: python fmea_poc.py [--drawing <PDF>] [--process <Excel>] [--samples <폴더>] [--output <파일명>]
      입력 파일 없으면 내장 데모 시나리오(프레스 브래킷)로 실행
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# ─── 내장 few-shot (AIAG-VDA 1st Ed. 표준 필드 기준) ─────────────────────────
BUILTIN_FEWSHOT = """
아래는 AIAG & VDA FMEA 1st Edition (2019) 기준 PFMEA 작성 예시 3건이다.

예시 1
- process_number: "10"
- process_step: "블랭킹"
- process_work_element: "소재 투입"
- function: "코일 소재를 설계 치수로 절단"
- special_characteristic: ""
- effect_end_user: "블랭크 형상 불량 → 후공정 치수 이탈 가능"
- S: 7
- effect_manufacturing: "금형 손상 / 후공정 자재 손실"
- failure_mode: "소재 방향 오투입"
- cause: "투입 방향 표시 미흡 / 작업자 오조작"
- O: 3
- prevention_controls: "소재 투입 방향 보조 지그 설치"
- detection_controls: "초물 치수 검사 (버니어 캘리퍼스)"
- D: 3, AP: "M"
- recommended_action: "소재 방향 감지 센서 추가 검토"

예시 2 (CC 특별특성)
- process_number: "40"
- process_step: "피어싱"
- process_work_element: "볼트홀 가공"
- function: "볼트홀 φ12 +0.1/-0.0mm 가공"
- special_characteristic: "CC"
- effect_end_user: "볼트 체결 불가 → 차량 안전사고"
- S: 9
- effect_manufacturing: "조립 라인 정지 / 납기 지연"
- failure_mode: "볼트홀 직경 과소 또는 과대"
- cause: "펀치 마모 / 다이 클리어런스 부적정"
- O: 2
- prevention_controls: "펀치 교체 기준 수립 (마모 한계 치수)"
- detection_controls: "핀 게이지 전수검사 (통과/불통과)"
- D: 2, AP: "H"
- recommended_action: "자동 핀 게이지 설비 도입 검토"

예시 3
- process_number: "60"
- process_step: "최종 검사"
- process_work_element: "치수 전수 검사"
- function: "설계 치수 내 합격품만 출하 승인"
- special_characteristic: ""
- effect_end_user: "고객 클레임 / 반품 발생"
- S: 8
- effect_manufacturing: "불량 재작업 비용 증가"
- failure_mode: "불량품 유출"
- cause: "검사 절차서 미준수 / 측정 기록 위변조"
- O: 2
- prevention_controls: "측정 결과 전산 입력 의무화"
- detection_controls: "측정 데이터 실시간 SPC 모니터링"
- D: 5, AP: "M"
- recommended_action: "측정 장비 RS-232 자동 수집 연동"
"""

DEMO_SCENARIO = """
부품명: 프런트 서브프레임 브래킷
부품번호: FSB-2024-001
재질: SPFC440 (고강도 냉연강판, 두께 2.0mm)
고객사: 현대자동차

주요 치수 및 공차:
- 전체 길이: 285 ± 0.5mm
- 볼트홀 직경: φ12 +0.1/-0.0mm (2개소) ← 특별특성(CC)
- 볼트홀 간격: 120 ± 0.3mm             ← 특별특성(CC)
- 플랜지 높이: 25 ± 0.5mm
- 표면조도: Ra 3.2 이하

공정 순서:
1. 블랭킹 — 코일 소재 절단 (프레스 250T)
2. 드로잉 1차 — 기본 형상 성형 (프레스 400T)
3. 드로잉 2차 — 플랜지 성형 (프레스 250T)
4. 피어싱 — 볼트홀 가공 (프레스 160T)
5. 트리밍 — 외형 정리 (프레스 160T)
6. 검사 — 치수 전수검사
"""

SYSTEM_PROMPT = """너는 자동차 부품 제조 공정의 PFMEA(공정 고장유형 및 영향분석) 전문가다.
AIAG & VDA FMEA 1st Edition(2019) 기준을 따른다.
RPN 방식이 아닌 AP(Action Priority: H/M/L) 방식을 사용한다.
출력은 반드시 유효한 JSON만 출력한다. 설명 텍스트나 마크다운 코드블록 외부에 아무것도 추가하지 않는다."""

PROMPT_TEMPLATE = """{sample_section}
## 분석 대상

{scenario}

---

위 부품과 공정에 대해 PFMEA를 작성해라.

규칙:
- 각 공정 단계마다 2~4개의 주요 고장유형 도출
- 특별특성은 special_characteristic 필드에 "CC" 또는 "SC" 기입 (없으면 빈 문자열)
- 고장 영향을 effect_end_user(최종 사용자)와 effect_manufacturing(제조 영향) 두 가지로 구분 기입
- process_number는 10, 20, 30... (같은 공정의 여러 항목은 동일 번호)
- AP 판정 기준:
  - H: S≥9이고 O≥3, 또는 S≥7이고 O≥4
  - M: S≥7이고 O≥2, 또는 S≥5이고 O≥6
  - L: 그 외
- responsibility, target_date, status, action_taken, revised_S/O/D/AP는 빈 문자열로 (담당자 직접 입력)

출력 형식:
```json
{{
  "part_name": "",
  "part_number": "",
  "customer": "",
  "rows": [
    {{
      "process_number": "10",
      "process_step": "공정명",
      "process_work_element": "작업 요소",
      "function": "기능/요구사항",
      "special_characteristic": "CC/SC/빈문자열",
      "effect_end_user": "최종 사용자 고장 영향",
      "S": 0,
      "effect_manufacturing": "제조 공정 영향",
      "failure_mode": "고장유형",
      "cause": "고장 원인",
      "O": 0,
      "prevention_controls": "예방 관리 방법(PC)",
      "detection_controls": "검출 관리 방법(DC)",
      "D": 0,
      "AP": "H/M/L",
      "recommended_action": "권고 조치",
      "responsibility": "",
      "target_date": "",
      "status": "",
      "action_taken": "",
      "revised_S": "",
      "revised_O": "",
      "revised_D": "",
      "revised_AP": ""
    }}
  ]
}}
```"""

# ─── 컬럼 정의 (AIAG-VDA 1st Ed. 표준 23컬럼) ───────────────────────────────

HEADERS = [
    # 구조 분석 (1~5)
    "번호", "공정 단계", "작업 요소", "기능/요구사항", "특별\n특성",
    # 고장 분석 (6~11)
    "고장 영향\n(최종 사용자)", "S", "고장 영향\n(제조)", "고장 유형", "고장 원인", "O",
    # 위험도 분석 (12~15)
    "예방 관리\n(PC)", "검출 관리\n(DC)", "D", "AP",
    # 최적화 (16~23)
    "권고 조치", "책임자/\n목표일", "현황", "조치 내용", "S*", "O*", "D*", "AP*",
]

FIELDS = [
    "process_number", "process_step", "process_work_element", "function", "special_characteristic",
    "effect_end_user", "S", "effect_manufacturing", "failure_mode", "cause", "O",
    "prevention_controls", "detection_controls", "D", "AP",
    "recommended_action", "_responsibility_date", "status", "action_taken",
    "revised_S", "revised_O", "revised_D", "revised_AP",
]

COL_WIDTHS = [5, 12, 14, 18, 6, 20, 4, 16, 18, 20, 4, 18, 18, 4, 5, 22, 16, 8, 18, 4, 4, 4, 5]

# 섹션 정의: (라벨, 시작 컬럼, 끝 컬럼, 색상코드)
SECTIONS = [
    ("구조 분석",   1,  5,  "2E75B6"),
    ("고장 분석",   6,  11, "ED7D31"),
    ("위험도 분석", 12, 15, "70AD47"),
    ("최적화",      16, 23, "7030A0"),
]

SECTION_LIGHT = {
    "2E75B6": "DEEAF1",
    "ED7D31": "FCE4D6",
    "70AD47": "E2EFDA",
    "7030A0": "EAE0F5",
}

AP_FILL = {"H": "FFCCCC", "M": "FFE4B5", "L": "CCFFCC"}
SC_FILL = {"CC": "C00000", "SC": "E46C0A"}


# ─── 입력 파싱 ────────────────────────────────────────────────────────────────

def parse_pdf_drawing(pdf_path: str) -> str:
    if not HAS_FITZ:
        return "(PyMuPDF 미설치 — pip install pymupdf)"
    doc = fitz.open(pdf_path)
    texts = [page.get_text() for page in doc]
    doc.close()
    raw = "\n".join(texts)[:3000]
    return raw or "(도면에서 텍스트 추출 실패 — 스캔 이미지일 수 있음)"


def parse_process_excel(excel_path: str) -> str:
    if not HAS_PANDAS:
        return "(pandas 미설치 — pip install pandas)"
    try:
        df = pd.read_excel(excel_path, header=None)
        return df.to_string(index=False, max_rows=50)[:2000]
    except Exception as e:
        return f"(Excel 파싱 오류: {e})"


def load_sample_fmea(sample_dir: str) -> str:
    samples = []
    for f in sorted(Path(sample_dir).glob("*.xlsx"))[:3]:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
                if any(c is not None for c in row):
                    rows.append(" | ".join(str(c) if c else "" for c in row))
            samples.append(f"[예시: {f.name}]\n" + "\n".join(rows[:12]))
            wb.close()
        except Exception:
            continue
    return "\n\n".join(samples)


# ─── Claude API 호출 ──────────────────────────────────────────────────────────

def generate_fmea(client: anthropic.Anthropic, scenario: str, samples: str) -> dict:
    all_samples = BUILTIN_FEWSHOT
    if samples:
        all_samples += f"\n\n## 추가 사내 예시\n\n{samples}"

    print("  Claude API 호출 중 (claude-sonnet-4-6)...", flush=True)
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        system=[{
            "type": "text",
            "text": SYSTEM_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }],
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": f"## PFMEA 작성 예시 (few-shot)\n\n{all_samples}\n\n---",
                    "cache_control": {"type": "ephemeral"},
                },
                {
                    "type": "text",
                    "text": PROMPT_TEMPLATE.format(sample_section="", scenario=scenario),
                },
            ],
        }],
        extra_headers={"anthropic-beta": "prompt-caching-2024-07-31"},
    )

    usage = msg.usage
    if getattr(usage, "cache_creation_input_tokens", 0):
        print(f"  캐시 저장: {usage.cache_creation_input_tokens:,} 토큰")
    if getattr(usage, "cache_read_input_tokens", 0):
        print(f"  캐시 히트: {usage.cache_read_input_tokens:,} 토큰 절감")

    raw = msg.content[0].text.strip()
    if "```json" in raw:
        raw = raw.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in raw:
        raw = raw.split("```", 1)[1].split("```", 1)[0].strip()

    return json.loads(raw)


# ─── Excel 출력 ───────────────────────────────────────────────────────────────

def write_excel(data: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PFMEA"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_cols = len(HEADERS)
    last_col = get_column_letter(total_cols)

    # Row 1: 타이틀
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = (f"PFMEA  —  {data.get('part_name','')}  "
               f"[{data.get('part_number','')}]  고객사: {data.get('customer','')}")
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2: 부제목
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (f"작성일: {datetime.now().strftime('%Y-%m-%d')}  |  "
               "기준: AIAG & VDA FMEA 1st Ed. (2019)  |  AP 방식  |  "
               "AI 초안 — 담당자 검토·승인 필수")
    c.font = Font(size=9, color="888888", italic=True)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Row 3: 섹션 헤더 (구조분석 | 고장분석 | 위험도분석 | 최적화)
    col_section_color = {}
    for label, c_start, c_end, color in SECTIONS:
        for ci in range(c_start, c_end + 1):
            col_section_color[ci] = color
        start_letter = get_column_letter(c_start)
        end_letter = get_column_letter(c_end)
        ws.merge_cells(f"{start_letter}3:{end_letter}3")
        cell = ws[f"{start_letter}3"]
        cell.value = label
        cell.font = Font(bold=True, size=9, color=color)
        cell.fill = PatternFill("solid", fgColor=SECTION_LIGHT[color])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 18

    # Row 4: 컬럼 헤더 (섹션별 색상)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=col_section_color.get(col, "2E75B6"))
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 36

    # Row 5+: 데이터
    for ri, row in enumerate(data.get("rows", []), 5):
        ap = str(row.get("AP", "L")).upper()
        sc = str(row.get("special_characteristic", "")).upper()

        for ci, field in enumerate(FIELDS, 1):
            # 책임자/목표일은 두 필드를 합쳐서 표시
            if field == "_responsibility_date":
                resp = row.get("responsibility", "")
                tgt = row.get("target_date", "")
                val = f"{resp}\n{tgt}".strip() if (resp or tgt) else ""
            else:
                val = row.get(field, "")

            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border

            if field == "special_characteristic" and sc in ("CC", "SC"):
                c.fill = PatternFill("solid", fgColor=SC_FILL[sc])
                c.font = Font(bold=True, color="FFFFFF", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field in ("S", "O", "D"):
                c.alignment = Alignment(horizontal="center", vertical="center")
                if field == "S" and isinstance(val, int) and val >= 9:
                    c.fill = PatternFill("solid", fgColor="FF6666")
            elif field == "AP":
                c.fill = PatternFill("solid", fgColor=AP_FILL.get(ap, "FFFFFF"))
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field in ("revised_S", "revised_O", "revised_D", "revised_AP"):
                c.fill = PatternFill("solid", fgColor="F2F2F2")
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[ri].height = 48

    # 열 너비
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    wb.save(output_path)


# ─── 메인 ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Phase 0 PoC — PFMEA 생성 (AIAG-VDA 1st Ed. 표준 양식)"
    )
    parser.add_argument("--drawing", help="도면 PDF 경로")
    parser.add_argument("--process", help="공정검토서 Excel 경로")
    parser.add_argument("--samples", help="기존 FMEA Excel 폴더 (few-shot용)")
    default_out = Path(__file__).parent / "output" / f"fmea_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    parser.add_argument("--output", default=str(default_out))
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY") or input("Anthropic API 키: ").strip()
    if not api_key:
        sys.exit("오류: API 키 필요")
    client = anthropic.Anthropic(api_key=api_key)

    print("\n[Phase 0 PoC] PFMEA 생성 시작 (AIAG-VDA 1st Ed. 표준)")
    print("=" * 55)

    if args.drawing:
        print(f"  도면 파싱: {args.drawing}")
        scenario = f"[도면]\n{parse_pdf_drawing(args.drawing)}"
        if args.process:
            print(f"  공정검토서 파싱: {args.process}")
            scenario += f"\n\n[공정검토서]\n{parse_process_excel(args.process)}"
    else:
        print("  입력 파일 없음 → 데모 시나리오 (프레스 브래킷, 6공정)")
        scenario = DEMO_SCENARIO

    samples = ""
    if args.samples:
        print(f"  샘플 로드: {args.samples}")
        samples = load_sample_fmea(args.samples)
        print(f"  → {samples.count('[예시:')}개 샘플 로드됨")

    try:
        data = generate_fmea(client, scenario, samples)
    except json.JSONDecodeError as e:
        sys.exit(f"JSON 파싱 오류: {e}")
    except anthropic.APIError as e:
        sys.exit(f"API 오류: {e}")

    rows = data.get("rows", [])
    h = sum(1 for r in rows if str(r.get("AP", "")).upper() == "H")
    m = sum(1 for r in rows if str(r.get("AP", "")).upper() == "M")
    l = sum(1 for r in rows if str(r.get("AP", "")).upper() == "L")
    cc = sum(1 for r in rows if str(r.get("special_characteristic", "")).upper() == "CC")
    sc = sum(1 for r in rows if str(r.get("special_characteristic", "")).upper() == "SC")

    print(f"\n  생성 완료: 총 {len(rows)}개 고장유형")
    print(f"    H (즉시 조치): {h}건  |  M (개선 권고): {m}건  |  L (저위험): {l}건")
    print(f"    특별특성 — CC: {cc}건, SC: {sc}건")

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(data, str(output_path))
    print(f"\n  Excel 저장: {output_path.resolve()}")

    print("\n" + "=" * 55)
    print("[품질/공정 담당자 검토 항목]")
    print("  1. 고장유형이 현장 실제 리스크를 반영하는가?")
    print("  2. S/O/D 점수가 현장 경험치와 일치하는가?")
    print("  3. 특별특성(CC/SC) 항목 누락 없는가?")
    print("  4. 최종 사용자 영향 / 제조 영향 구분이 적절한가?")
    print("  5. AP=H 항목 권고 조치가 실효성 있는가?")
    print("  → 합격: Phase 1 진행 | 불합격: 프롬프트 수정 후 재실행")
    print("=" * 55)


if __name__ == "__main__":
    main()
