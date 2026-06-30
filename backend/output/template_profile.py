"""
고객사 양식(.xlsx/.xlsm) 업로드 → 출력 프로파일 자동 생성.

업로드된 FMEA 양식을 파서로 분석해 (시트·헤더·필드→컬럼) 매핑을 자동 추출하고,
default 프로파일을 복제한 뒤 FMEA만 template 렌더러로 오버라이드한다.
→ FMEA는 고객사 양식 그대로, 나머지 3종(CP·작업표준서·자주검사)은 기존 config 양식 유지.

사람이 컬럼 매핑을 손으로 정의할 필요가 없다(파서가 자동 감지).
"""

from __future__ import annotations

import copy
import json
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from ..preprocessing.excel_parser import parse_fmea_excel, build_value_map, _normalize_text, _compact
from .profiles import load_profile, PROFILE_DIR

TEMPLATE_DIR = Path(__file__).parent / "templates"

# 파서 필드명 → 에이전트 데이터 필드명
_FIELD_TRANSLATE = {"rpn_ap": "RPN", "effect": "effect_end_user"}

# 메타 셀 라벨 키워드 → 데이터 필드
_META_LABELS = {
    "part_number": ["품번", "품 번", "부품번호", "part no", "part number"],
    "part_name": ["품명", "품 명", "부품명", "part name", "part description"],
    "vehicle_model": ["적용차종", "차종", "vehicle", "engine program"],
    "fmea_no": ["fmea no", "fmea번호", "fmea 번호"],
}


def _slug(name: str) -> str:
    s = re.sub(r"[^0-9A-Za-z가-힣]+", "_", (name or "").strip()).strip("_")
    return s.lower() or "customer"


def _translate_columns(col_map: dict) -> dict:
    out: dict[str, int] = {}
    for f, c in col_map.items():
        out[_FIELD_TRANSLATE.get(f, f)] = c
    return out


def _detect_data_start(rows: list[dict], header_row: int) -> int:
    """헤더 아래 첫 '실데이터' 행(심각도 S가 숫자) 위치. 못 찾으면 header_row+1."""
    for i, row in enumerate(rows):
        if row.get("S") is not None:
            return header_row + 1 + i
    return header_row + 1


def _detect_meta_cells(file_path: str, sheet_name: str, header_row: int) -> dict:
    """헤더 영역에서 품번/품명/차종 라벨을 찾아 우측 칸을 메타 타겟으로 매핑."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        vm = build_value_map(ws)
        max_col = max((c for _, c in vm), default=1)
        meta: dict[str, str] = {}
        used_fields: set[str] = set()
        for r in range(1, header_row + 1):
            for c in range(1, max_col):
                val = _normalize_text(vm.get((r, c), "")).lower()
                if not val:
                    continue
                val_c = _compact(val)
                for field, kws in _META_LABELS.items():
                    if field in used_fields:
                        continue
                    if any(kw in val or _compact(kw) in val_c for kw in kws):
                        ref = f"{get_column_letter(c + 1)}{r}"
                        meta[ref] = field
                        used_fields.add(field)
                        break
        wb.close()
        return meta
    except Exception:
        return {}


def build_profile(file_path: str, customer: str) -> dict:
    """업로드된 양식 파일 → 프로파일 dict 생성 (저장하지 않음). 분석 요약 포함."""
    parsed = parse_fmea_excel(file_path)
    if "error" in parsed:
        raise ValueError(parsed["error"])

    col_map = _translate_columns(parsed["column_map"])
    if not col_map or "failure_mode" not in col_map:
        raise ValueError(
            "FMEA 양식으로 인식하지 못했습니다. (고장유형/심각도 등 컬럼 미감지) "
            f"감지된 컬럼: {list(col_map.keys())}"
        )

    header_row = parsed["header_row"]
    sheet_name = parsed["sheet_name"]
    data_start = _detect_data_start(parsed.get("rows", []), header_row)
    meta_cells = _detect_meta_cells(file_path, sheet_name, header_row)

    # default 프로파일 복제 → FMEA만 template 으로 오버라이드
    profile = copy.deepcopy(load_profile(None))
    profile["company"] = customer
    profile["aliases"] = [customer, _slug(customer)]
    profile.setdefault("fmea", {})["renderer"] = "template"
    profile["template"] = {
        "fmea": {
            "path": f"templates/{Path(file_path).name}",
            "sheet": sheet_name,
            "data_start_row": data_start,
            "columns": col_map,
            "meta_cells": meta_cells,
        }
    }
    profile["_analysis"] = {
        "sheet": sheet_name,
        "header_row": header_row,
        "data_start_row": data_start,
        "detected_fields": list(col_map.keys()),
        "meta_cells": meta_cells,
    }
    return profile


def save_uploaded_template(file_bytes: bytes, filename: str, customer: str) -> dict:
    """업로드 파일을 저장하고 프로파일 생성·기록. 분석 요약 반환."""
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    ext = Path(filename).suffix.lower() or ".xlsx"
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError("Excel 양식 파일(.xlsx 또는 .xlsm)만 업로드할 수 있습니다.")

    safe_name = f"{_slug(customer)}{ext}"
    dest = TEMPLATE_DIR / safe_name
    dest.write_bytes(file_bytes)

    try:
        profile = build_profile(str(dest), customer)
    except Exception:
        dest.unlink(missing_ok=True)
        raise

    # 템플릿 path를 실제 저장 파일명으로 교정
    profile["template"]["fmea"]["path"] = f"templates/{safe_name}"

    profile_path = PROFILE_DIR / f"{_slug(customer)}.json"
    profile_path.write_text(
        json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # 프로파일 캐시 무효화 (다음 load_profile 에서 반영)
    load_profile.cache_clear() if hasattr(load_profile, "cache_clear") else None
    from .profiles import _load_all
    _load_all.cache_clear()

    analysis = profile.pop("_analysis", {})
    return {
        "customer": customer,
        "template_file": safe_name,
        "profile_file": profile_path.name,
        **analysis,
    }


def list_templates() -> list[dict]:
    """업로드된 고객사 양식 프로파일 목록."""
    out = []
    for jf in PROFILE_DIR.glob("*.json"):
        try:
            data = json.loads(jf.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        if "template" not in data:  # default.json 등 일반 프로파일 제외
            continue
        tpl = (data.get("template") or {}).get("fmea", {})
        out.append({
            "customer": data.get("company", jf.stem),
            "profile_file": jf.name,
            "template_file": Path(tpl.get("path", "")).name,
            "sheet": tpl.get("sheet", ""),
            "detected_fields": list((tpl.get("columns") or {}).keys()),
        })
    return out


def delete_template(customer: str) -> bool:
    """고객사 양식 프로파일 + 템플릿 파일 삭제."""
    profile_path = PROFILE_DIR / f"{_slug(customer)}.json"
    removed = False
    if profile_path.exists():
        try:
            data = json.loads(profile_path.read_text(encoding="utf-8"))
            tpl_name = Path((data.get("template") or {}).get("fmea", {}).get("path", "")).name
            if tpl_name:
                (TEMPLATE_DIR / tpl_name).unlink(missing_ok=True)
        except Exception:
            pass
        profile_path.unlink(missing_ok=True)
        removed = True
    from .profiles import _load_all
    _load_all.cache_clear()
    return removed
