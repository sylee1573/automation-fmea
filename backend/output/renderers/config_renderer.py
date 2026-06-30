"""
Config 렌더러 — 회사 프로파일(JSON)의 컬럼·라벨·색 설정대로 Excel을 코드로 그린다.

기존 excel_generator.py의 그리기 로직을 일반화한 것.
프로파일 doc 설정 형식:
  {
    "title": "PFMEA",
    "standard": "AIAG FMEA 4th Ed. (RPN)",   # 부제목에 표기
    "header_color": "2E75B6",
    "sections": [["구조 분석", 1, 5, "2E75B6"], ...],   # 선택(밴드 행). 없으면 단일 헤더행
    "columns": [{"label": "번호", "field": "process_number", "width": 5}, ...]
  }
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 공통 스타일 상수 ──────────────────────────────────────────────────────────

_SECTION_LIGHT = {
    "2E75B6": "DEEAF1", "ED7D31": "FCE4D6", "70AD47": "E2EFDA", "7030A0": "EAE0F5",
}
_RPN_FILL = [(100, "FFCCCC"), (40, "FFE4B5"), (0, "CCFFCC")]  # 임계값 내림차순
_SC_FILL = {"CC": "C00000", "SC": "E46C0A"}

_CENTER_FIELDS = {
    "process_number", "step_number", "characteristic_type", "sample_size",
    "S", "O", "D", "RPN", "revised_S", "revised_O", "revised_D", "revised_RPN",
}
_REVISED_FIELDS = {"revised_S", "revised_O", "revised_D", "revised_RPN"}


def _thin_border() -> Border:
    side = Side(style="thin", color="BBBBBB")
    return Border(left=side, right=side, top=side, bottom=side)


def _title_row(ws, value: str, total_cols: int, row: int = 1):
    last = get_column_letter(total_cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = value
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def _subtitle_row(ws, value: str, total_cols: int, row: int = 2):
    last = get_column_letter(total_cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = value
    c.font = Font(size=9, color="888888", italic=True)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 16


def _rpn_fill(value) -> str | None:
    try:
        v = int(value)
    except (TypeError, ValueError):
        return None
    for threshold, color in _RPN_FILL:
        if v >= threshold:
            return color
    return None


def _doc_title(doc_cfg: dict, data: dict) -> str:
    pn = data.get("part_number", "")
    nm = data.get("part_name", "")
    cust = data.get("customer", "")
    return f"{doc_cfg.get('title', '')}  —  {nm}  [{pn}]  고객사: {cust}"


def render(doc_type: str, data: dict, profile: dict, output_path: str) -> Path:
    """프로파일 doc 설정대로 단일 문서 Excel 생성."""
    doc_cfg = profile[doc_type]
    columns = doc_cfg["columns"]
    sections = doc_cfg.get("sections")
    base_color = doc_cfg.get("header_color", "2E75B6")
    total_cols = len(columns)
    border = _thin_border()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doc_cfg.get("title", doc_type)[:31]

    _title_row(ws, _doc_title(doc_cfg, data), total_cols, row=1)
    standard = doc_cfg.get("standard", "")
    subtitle = f"작성일: {datetime.now().strftime('%Y-%m-%d')}"
    if standard:
        subtitle += f"  |  기준: {standard}"
    subtitle += "  |  AI 초안 — 담당자 검토·승인 필수"
    _subtitle_row(ws, subtitle, total_cols, row=2)

    # ── 헤더 (선택적 섹션 밴드 + 컬럼 헤더) ────────────────────────────────────
    col_section_color: dict[int, str] = {}
    if sections:
        for label, c_start, c_end, color in sections:
            for ci in range(c_start, c_end + 1):
                col_section_color[ci] = color
            ws.merge_cells(f"{get_column_letter(c_start)}3:{get_column_letter(c_end)}3")
            cell = ws[f"{get_column_letter(c_start)}3"]
            cell.value = label
            cell.font = Font(bold=True, size=9, color=color)
            cell.fill = PatternFill("solid", fgColor=_SECTION_LIGHT.get(color, "F2F2F2"))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[3].height = 18
        header_row = 4
    else:
        header_row = 3

    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=header_row, column=ci, value=col["label"])
        c.fill = PatternFill("solid", fgColor=col_section_color.get(ci, base_color))
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 36

    # ── 데이터 행 ──────────────────────────────────────────────────────────────
    data_start = header_row + 1
    for ri, row in enumerate(data.get("rows", []), data_start):
        sc = str(row.get("special_characteristic", "")).upper()
        for ci, col in enumerate(columns, 1):
            field = col["field"]
            if field == "_resp_date":
                resp = row.get("responsibility", "")
                tgt = row.get("target_date", "")
                val = f"{resp}\n{tgt}".strip() if (resp or tgt) else ""
            else:
                val = row.get(field, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border

            if field == "special_characteristic" and sc in ("CC", "SC"):
                c.fill = PatternFill("solid", fgColor=_SC_FILL[sc])
                c.font = Font(bold=True, color="FFFFFF", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field == "RPN":
                fill = _rpn_fill(val)
                if fill:
                    c.fill = PatternFill("solid", fgColor=fill)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field == "S":
                c.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, int) and val >= 9:
                    c.fill = PatternFill("solid", fgColor="FF6666")
            elif field in _REVISED_FIELDS:
                c.fill = PatternFill("solid", fgColor="F2F2F2")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field in _CENTER_FIELDS:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = doc_cfg.get("row_height", 44)

    for ci, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col.get("width", 12)
    ws.freeze_panes = f"A{data_start}"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return Path(output_path)
