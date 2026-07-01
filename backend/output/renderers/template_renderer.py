"""
Template 렌더러 — 고객사가 업로드한 실제 Excel 양식을 틀로 사용해 데이터를 채운다.

config 렌더러가 양식을 '코드로 그리는' 반면, 이 렌더러는 고객사 원본 양식 파일
(헤더·로고·서식·컬럼폭 그대로)을 열어 데이터 행만 채워 넣는다.

프로파일 형식 (profiles/*.json):
  {
    "renderer": "template",
    "template": {
      "fmea": {
        "path": "templates/mtk.xlsm",     # backend/output 기준 상대경로 또는 절대경로
        "sheet": "PFMEA",
        "data_start_row": 11,
        "columns": {"process_step": 2, "failure_mode": 3, "S": 5, ...},  # field→열번호
        "meta_cells": {"C7": "part_number", "C8": "part_name"}            # 선택
      }
    }
  }
"""

from __future__ import annotations

import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import RowBreak, ColBreak

_TEMPLATE_BASE = Path(__file__).parent.parent  # backend/output

# 점수 칸(S·O·D·RPN 등)은 폭이 좁으므로 줄바꿈 대신 자동 축소로 3자리도 표시.
_NUMERIC_FIELDS = {
    "S", "O", "D", "RPN",
    "revised_S", "revised_O", "revised_D", "revised_RPN",
}


def _resolve(path_str: str) -> Path:
    p = Path(path_str)
    return p if p.is_absolute() else (_TEMPLATE_BASE / p)


def _thin_border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _col_width(ws, col: int) -> float:
    w = ws.column_dimensions[get_column_letter(col)].width
    return w if w else 8.43


def _est_lines(text: str, width: float) -> int:
    """열 폭 대비 필요한 줄 수 근사 (한글은 폭 2로 계산, 명시적 개행 포함)."""
    cap = max(width - 1.0, 1.0)
    total = 0
    for line in str(text).split("\n"):
        units = sum(2 if ord(ch) > 0x2E80 else 1 for ch in line)
        total += max(1, math.ceil(units / cap))
    return total


def render(doc_type: str, data: dict, profile: dict, output_path: str) -> Path:
    """고객사 양식 템플릿에 data["rows"]를 채워 저장."""
    tpl_cfg = (profile.get("template") or {}).get(doc_type)
    if not tpl_cfg:
        # 이 문서 타입에 템플릿이 없으면 config 렌더러로 폴백
        from . import config_renderer
        return config_renderer.render(doc_type, data, profile, output_path)

    tpl_path = _resolve(tpl_cfg["path"])
    if not tpl_path.exists():
        raise FileNotFoundError(f"양식 템플릿 없음: {tpl_path}")

    keep_vba = tpl_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(tpl_path, keep_vba=keep_vba)

    sheet = tpl_cfg.get("sheet")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    columns: dict = {str(k): int(v) for k, v in tpl_cfg["columns"].items()}
    start = int(tpl_cfg.get("data_start_row", 2))
    form_right = max(max(columns.values(), default=0), ws.max_column)  # 양식 전체 폭(U 등)
    border = _thin_border()

    # 데이터 바로 위 2행(헤더)의 가로 병합 = 한 필드가 여러 열을 차지(예: 원인 G:H,
    # 현관리 J:K, 권고조치 N:O). 같은 열 스팬을 데이터 행에도 적용해 세로쓰기 방지.
    data_spans = [
        (rng.min_col, rng.max_col)
        for rng in ws.merged_cells.ranges
        if rng.min_col < rng.max_col
        and rng.min_row == start - 2 and rng.max_row == start - 1
    ]
    span_width = {a: sum(_col_width(ws, c) for c in range(a, b + 1)) for a, b in data_spans}

    # 데이터 영역(헤더 아래) 병합 해제 + 기존 내용 지우기
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start:
            ws.unmerge_cells(str(rng))
    for r in range(start, (ws.max_row or start) + 1):
        for c in range(1, form_right + 1):
            ws.cell(row=r, column=c).value = None

    # 데이터 채우기
    rows = data.get("rows", [])
    for i, row in enumerate(rows):
        r = start + i

        # 양식 전체 폭에 격자 테두리 (병합 해제 후 내부 테두리 누락 방지)
        for c in range(2, form_right + 1):
            ws.cell(row=r, column=c).border = border

        max_lines = 1
        for field, col in columns.items():
            val = row.get(field, "")
            # MTK처럼 예방/검출 관리가 한 칸인 경우: 예방이 비면 검출로 대체
            if field == "prevention_controls" and not val:
                val = row.get("detection_controls", "") or ""
            # 빈 값은 건너뜀 — 같은 컬럼에 여러 필드가 매핑된 경우(process_step/function)
            # 빈 필드가 이미 채운 값을 덮어쓰지 않도록 보호
            if val in (None, ""):
                continue
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            if field in _NUMERIC_FIELDS:
                # 숫자 칸: 줄바꿈 대신 자동 축소(##·3자리 방지), 가운데 정렬
                cell.alignment = Alignment(
                    vertical="center", horizontal="center", shrink_to_fit=True
                )
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                eff_w = span_width.get(col, _col_width(ws, col))
                max_lines = max(max_lines, _est_lines(val, eff_w))

        # 병합 스팬을 데이터 행에 적용 (값은 이미 왼쪽 열에 기록됨)
        for a, b in data_spans:
            ws.merge_cells(start_row=r, start_column=a, end_row=r, end_column=b)

        # 내용 길이에 맞춰 행 높이 (한 줄 ≈ 15pt, 여백 6)
        ws.row_dimensions[r].height = min(150, max(30, max_lines * 15 + 6))

    # 데이터 아래 잔여 템플릿 행 삭제 → 빈 2~7페이지·테두리 잔재 제거
    last = start + len(rows) - 1
    if rows and ws.max_row > last:
        ws.delete_rows(last + 1, ws.max_row - last)
    # 수동 페이지 나눔 제거 + 인쇄영역을 실제 데이터 범위로
    ws.row_breaks = RowBreak()
    ws.col_breaks = ColBreak()
    if rows:
        ws.print_area = f"B2:{get_column_letter(form_right)}{last}"

    # 메타 셀 채우기 (품번/품명/차종 등)
    for cell_ref, field in (tpl_cfg.get("meta_cells") or {}).items():
        v = data.get(field, "")
        if v:
            try:
                ws[cell_ref] = v
            except Exception:
                pass

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return Path(output_path)
