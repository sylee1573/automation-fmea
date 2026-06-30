#!/usr/bin/env python3
"""
Phase 2: FMEA Excel 파서

기존 FMEA Excel 파일을 표준 스키마로 정규화.
셀 병합 해제, 버전별 컬럼 자동 매핑, 데이터 추출.

실행: python excel_parser.py --file <FMEA.xlsx> [--output parsed.json]
"""

import argparse
import json
from pathlib import Path
from typing import Optional

import openpyxl

# ─── 컬럼 매핑 테이블 (다양한 FMEA 버전 대응) ───────────────────────────────
COLUMN_MAP: dict[str, list[str]] = {
    "process_step": [
        "공정 단계", "공정단계", "공정명", "공정", "process step",
        "process item", "단계", "step",
    ],
    "process_work_element": [
        "작업 요소", "작업요소", "work element", "작업내용", "작업 내용",
    ],
    "function": [
        "기능", "기능/요구사항", "요구사항", "function", "requirements",
        "기능 및 요구사항",
    ],
    "failure_mode": [
        "고장유형", "고장 유형", "불량유형", "불량 유형",
        "failure mode", "fm", "고장모드", "고장형태", "잠재적 고장형태",
    ],
    "effect": [
        "고장 영향", "고장영향", "영향", "effect",
        "고객영향", "failure effect",
    ],
    "effect_end_user": [
        "고장 영향(최종 사용자)", "고장 영향\n(최종 사용자)",
        "최종 사용자 영향", "effect end user",
    ],
    "effect_manufacturing": [
        "고장 영향(제조)", "고장 영향\n(제조)",
        "제조 영향", "manufacturing effect",
    ],
    "cause": [
        "고장 원인", "고장원인", "원인", "cause", "failure cause",
    ],
    "S": ["S", "심각도", "심각", "severity"],
    "O": ["O", "발생도", "발생", "occurrence"],
    "D": ["D", "검출도", "검출", "detection"],
    "rpn_ap": ["RPN", "AP", "action priority", "위험도"],
    "prevention_controls": [
        "예방관리", "예방 관리", "관리방법", "현 관리방법(예방)",
        "prevention controls", "예방조치", "예방 관리\n(PC)", "pc",
        "현 공정관리", "공정관리",
    ],
    "detection_controls": [
        "검출관리", "검출 관리", "현 관리방법(검출)",
        "detection controls", "검출방법", "검출 관리\n(DC)", "dc",
    ],
    "recommended_action": [
        "권고조치", "권고 조치", "개선조치", "개선안",
        "recommended action", "권고 사항",
    ],
    "special_characteristic": [
        "특별특성", "특별 특성", "cc/sc", "special characteristic",
        "특별\n특성",
    ],
    "responsibility": [
        "책임자", "담당자", "책임자/목표일", "책임자/\n목표일",
        "responsibility",
    ],
}


def _normalize_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ")


def _compact(s: str) -> str:
    """공백·점·슬래시 제거 — 세로텍스트(심\\n각\\n도) / 점표기(R.P.N.) 헤더 매칭용"""
    return s.replace(" ", "").replace(".", "").replace("/", "")


def _header_matches(header_val: str, keyword: str) -> bool:
    """헤더 셀이 키워드와 매칭되는지 (일반 + 공백/점 제거 후 비교)"""
    h, k = header_val.lower(), keyword.lower()
    if k in h or h in k:
        return True
    hc, kc = _compact(h), _compact(k)
    return bool(hc) and bool(kc) and (kc in hc or hc in kc)


def build_value_map(ws) -> dict[tuple, any]:
    """병합 셀 포함 전체 셀 값 맵 (row, col) → value"""
    value_map: dict[tuple, any] = {}

    # 병합 셀: 상단좌측 값으로 채움
    for merged in ws.merged_cells.ranges:
        top_left = ws.cell(merged.min_row, merged.min_col).value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                value_map[(r, c)] = top_left

    # 일반 셀
    for row in ws.iter_rows():
        for cell in row:
            key = (cell.row, cell.column)
            if key not in value_map:
                value_map[key] = cell.value

    return value_map


def detect_header_row(value_map: dict, max_scan: int = 15) -> int:
    """FMEA 표준 필드가 가장 많이(서로 다른 종류로) 매칭되는 행 번호 반환.

    서로 다른 필드 수로 점수를 매겨, 동일 제목이 여러 칸에 병합된 타이틀 행
    (예: MTK 양식 R2 '잠재적 고장형태 및 영향분석')이 헤더로 오인되지 않게 한다.
    """
    max_col = max((c for _, c in value_map), default=1)

    best_row, best_score = 1, 0
    for r in range(1, max_scan + 1):
        matched_fields = set()
        for c in range(1, max_col + 1):
            val = _normalize_text(value_map.get((r, c), ""))
            if not val:
                continue
            for field, keywords in COLUMN_MAP.items():
                if any(_header_matches(val, kw) for kw in keywords):
                    matched_fields.add(field)
        score = len(matched_fields)
        if score > best_score:
            best_score, best_row = score, r

    return best_row


def map_columns(value_map: dict, header_row: int) -> dict[str, int]:
    """헤더 행에서 표준 필드명 → 컬럼 인덱스 매핑"""
    max_col = max((c for _, c in value_map), default=1)
    col_to_field: dict[str, int] = {}

    for c in range(1, max_col + 1):
        header_val = _normalize_text(value_map.get((header_row, c), ""))
        if not header_val:
            continue
        for field, keywords in COLUMN_MAP.items():
            for kw in keywords:
                if _header_matches(header_val, kw):
                    if field not in col_to_field:  # 첫 매칭만 사용
                        col_to_field[field] = c
                    break

    return col_to_field


def extract_rows(
    value_map: dict,
    header_row: int,
    col_map: dict[str, int],
    max_row: int,
) -> list[dict]:
    rows = []
    for r in range(header_row + 1, max_row + 1):
        # 완전 빈 행 건너뜀
        row_vals = [value_map.get((r, c)) for c in range(1, max(col_map.values(), default=1) + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue

        record: dict[str, any] = {}
        for field, c in col_map.items():
            raw = value_map.get((r, c))
            # S/O/D/RPN은 숫자로 변환 시도
            if field in ("S", "O", "D"):
                try:
                    record[field] = int(float(raw)) if raw is not None else None
                except (ValueError, TypeError):
                    record[field] = None
            else:
                record[field] = _normalize_text(raw) or None

        # effect 통합 필드 → effect_end_user 로 폴백
        if "effect_end_user" not in record and "effect" in record:
            record["effect_end_user"] = record.pop("effect")

        rows.append(record)

    return rows


def _select_fmea_sheet(wb):
    """워크북에서 FMEA 컬럼이 가장 많이 잡히는 시트 선택.
    반환: (worksheet, value_map, header_row, col_map)"""
    best = None  # (score, ws, value_map, header_row, col_map)
    for ws in wb.worksheets:
        value_map = build_value_map(ws)
        header_row = detect_header_row(value_map)
        col_map = map_columns(value_map, header_row)
        score = len(col_map)
        if best is None or score > best[0]:
            best = (score, ws, value_map, header_row, col_map)
    _, ws, value_map, header_row, col_map = best
    return ws, value_map, header_row, col_map


def parse_fmea_excel(file_path: str) -> dict:
    """
    FMEA Excel 파일 파싱.
    Returns:
        file_path, sheet_name, header_row, column_map, rows, warnings
    """
    warnings: list[str] = []
    path = Path(file_path)

    if not path.exists():
        return {"error": f"파일 없음: {file_path}"}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {"error": f"파일 열기 실패: {e}"}

    # 여러 시트 중 FMEA 헤더가 가장 잘 매칭되는 시트 선택
    # (MTK 양식처럼 active 시트가 PFD인 경우 대비)
    ws, value_map, header_row, col_map = _select_fmea_sheet(wb)
    max_row = ws.max_row or 1

    # 필수 컬럼 누락 경고
    required = ["failure_mode", "S", "O", "D"]
    for f in required:
        if f not in col_map:
            warnings.append(f"필수 컬럼 미감지: {f}")

    rows = extract_rows(value_map, header_row, col_map, max_row)
    wb.close()

    # RPN vs AP 방식 감지
    mode = "AP" if "rpn_ap" in col_map and any(
        str(r.get("rpn_ap", "")).upper() in ("H", "M", "L") for r in rows
    ) else "RPN"

    return {
        "file_path": str(path.resolve()),
        "sheet_name": ws.title,
        "header_row": header_row,
        "column_map": col_map,
        "fmea_mode": mode,
        "row_count": len(rows),
        "rows": rows,
        "warnings": warnings,
    }


# ─── PFD(공정흐름도) 컬럼 매핑 ───────────────────────────────────────────────

PFD_COLUMN_MAP: dict[str, list[str]] = {
    "process_no": [
        "공정번호", "공정 번호", "no", "no.", "번호", "process no", "op no",
        "공정 no", "공정no",
    ],
    "process_name": [
        "공정명", "공정 명", "공정이름", "process name", "operation", "공정",
        "작업명", "공정 내용",
    ],
    "process_type": [
        "공정유형", "공정 유형", "유형", "type", "공정종류", "종류",
    ],
    "machine": [
        "기계", "설비", "기계/설비", "장비", "machine", "equipment",
        "기계설비", "사용설비", "치구", "금형",
    ],
    "function": [
        "기능", "공정기능", "공정 기능", "목적", "function", "purpose",
        "기능/목적",
    ],
    "input": [
        "투입", "입력", "투입물", "투입소재", "input", "소재", "투입 부품",
        "투입재료",
    ],
    "output": [
        "산출", "출력", "산출물", "output", "제품", "완성품",
    ],
    "special_char": [
        "특별특성", "특별 특성", "cc", "sc", "cc/sc", "특성기호",
        "특별특성기호", "special characteristic",
    ],
    "remark": [
        "비고", "특기사항", "remark", "note", "notes", "참고",
    ],
}


def _detect_pfd_header(value_map: dict, max_scan: int = 15) -> int:
    """PFD 헤더 행 자동 감지"""
    all_keywords = {kw.lower() for kws in PFD_COLUMN_MAP.values() for kw in kws}
    max_col = max((c for _, c in value_map), default=1)
    best_row, best_score = 1, 0
    for r in range(1, max_scan + 1):
        score = sum(
            1 for c in range(1, max_col + 1)
            if any(kw in _normalize_text(value_map.get((r, c), "")).lower()
                   for kw in all_keywords)
        )
        if score > best_score:
            best_score, best_row = score, r
    return best_row


def _map_pfd_columns(value_map: dict, header_row: int) -> dict[str, int]:
    """PFD 헤더에서 필드명 → 컬럼 인덱스 매핑"""
    max_col = max((c for _, c in value_map), default=1)
    col_to_field: dict[str, int] = {}
    for c in range(1, max_col + 1):
        header_val = _normalize_text(value_map.get((header_row, c), "")).lower()
        if not header_val:
            continue
        for field, keywords in PFD_COLUMN_MAP.items():
            if field in col_to_field:
                continue
            if any(kw.lower() in header_val or header_val in kw.lower()
                   for kw in keywords):
                col_to_field[field] = c
    return col_to_field


def parse_pfd_excel(file_path: str) -> dict:
    """
    PFD(공정흐름도) Excel 파싱.
    Returns: steps 리스트 (공정번호/공정명/설비/기능/특별특성 등)
    """
    path = Path(file_path)
    if not path.exists():
        return {"error": f"파일 없음: {file_path}"}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {"error": f"파일 열기 실패: {e}"}

    ws = wb.active
    value_map = build_value_map(ws)
    max_row = ws.max_row or 1

    header_row = _detect_pfd_header(value_map)
    col_map = _map_pfd_columns(value_map, header_row)

    steps = []
    for r in range(header_row + 1, max_row + 1):
        row_vals = [value_map.get((r, c)) for c in range(1, max(col_map.values(), default=1) + 1)]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        step: dict[str, str] = {}
        for field, c in col_map.items():
            step[field] = _normalize_text(value_map.get((r, c)) or "")
        # 공정명이 없으면 건너뜀
        if not step.get("process_name") and not step.get("process_no"):
            continue
        steps.append(step)

    wb.close()

    detected = list(col_map.keys())
    warnings = []
    for required in ("process_name", "machine", "function"):
        if required not in col_map:
            warnings.append(f"PFD 컬럼 미감지: {required}")

    return {
        "file_path": str(path.resolve()),
        "sheet_name": ws.title,
        "header_row": header_row,
        "detected_columns": detected,
        "step_count": len(steps),
        "steps": steps,
        "warnings": warnings,
    }


def parse_pfd(content: bytes) -> str:
    """
    bytes로 받은 PFD Excel을 PFMEA 에이전트 주입용 텍스트로 변환.
    공정번호·공정명·설비·기능·특별특성을 구조화된 텍스트로 반환.
    """
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        result = parse_pfd_excel(tmp_path)
        if "error" in result:
            return f"(PFD 파싱 오류: {result['error']})"

        lines = ["[공정흐름도 (PFD)]"]
        lines.append(f"총 {result['step_count']}개 공정 감지")
        lines.append("")

        for i, step in enumerate(result["steps"], 1):
            no = step.get("process_no", str(i))
            name = step.get("process_name", "(미기재)")
            parts = [f"[{no}] {name}"]
            if step.get("machine"):
                parts.append(f"  설비: {step['machine']}")
            if step.get("function"):
                parts.append(f"  기능: {step['function']}")
            if step.get("input"):
                parts.append(f"  투입: {step['input']}")
            if step.get("output"):
                parts.append(f"  산출: {step['output']}")
            if step.get("special_char"):
                parts.append(f"  특별특성: {step['special_char']}")
            if step.get("remark"):
                parts.append(f"  비고: {step['remark']}")
            lines.extend(parts)
            lines.append("")

        if result["warnings"]:
            lines.append("※ 주의: " + " / ".join(result["warnings"]))

        return "\n".join(lines)
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def parse_process_sheet(content: bytes) -> str:
    """
    bytes로 받은 공정검토서 Excel을 텍스트로 변환 (main.py FastAPI 업로드 어댑터).
    에이전트 프롬프트 주입용 텍스트 반환.
    """
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active
        lines = []
        for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            line = " | ".join(c for c in cells if c)
            if line:
                lines.append(line)
        wb.close()
        return "\n".join(lines[:60])
    except Exception as e:
        return f"(공정검토서 파싱 오류: {e})"
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def main():
    parser = argparse.ArgumentParser(description="Phase 2 — FMEA Excel 파서")
    parser.add_argument("--file", required=True, help="파싱할 FMEA Excel 파일")
    parser.add_argument("--output", default="", help="JSON 출력 파일 (기본: 콘솔 출력)")
    args = parser.parse_args()

    result = parse_fmea_excel(args.file)

    if "error" in result:
        print(f"오류: {result['error']}")
        return

    print(f"[excel_parser] {Path(args.file).name}")
    print(f"  헤더 행: {result['header_row']}행")
    print(f"  감지 컬럼: {list(result['column_map'].keys())}")
    print(f"  FMEA 방식: {result['fmea_mode']}")
    print(f"  데이터 행: {result['row_count']}개")
    if result["warnings"]:
        for w in result["warnings"]:
            print(f"  ⚠ {w}")

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"  저장: {args.output}")
    else:
        print(json.dumps(result["rows"][:3], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
