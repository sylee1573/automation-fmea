#!/usr/bin/env python3
"""
Phase 1: FMEA Excel 파일 스캐너

공유 서버 폴더를 재귀 스캔하여 Excel 파일 목록표 CSV 출력.
실행: python excel_scanner.py --dir <폴더경로> [--output 목록표.csv]

출력 CSV 컬럼:
  파일명, 경로, 파일크기_KB, 수정일, 시트수, 데이터행수_추정,
  FMEA_판별(O/X), S_O_D_컬럼(있음/없음), AP_RPN_방식,
  공장명_추정, 공정유형_추정, 고객사_추정, 연도_추정, 오류
"""

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path

import openpyxl

FACTORY_KEYWORDS = ["프레스", "용접", "열처리", "가공", "도장", "조립"]
PROCESS_KEYWORDS = ["드로잉", "피어싱", "블랭킹", "트리밍", "MIG", "TIG", "스폿", "침탄", "고주파", "선반", "MCT", "밀링", "정전", "분체", "조립"]
CUSTOMER_KEYWORDS = ["현대", "기아", "한국GM", "GM", "르노", "삼성", "쌍용", "현기"]


def detect_fmea_columns(ws) -> dict:
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                headers.append(cell.strip())

    header_str = " ".join(headers).upper()

    has_severity = any(k in header_str for k in ["심각도", "SEVERITY", "중요도", "심각"])
    has_occurrence = any(k in header_str for k in ["발생도", "OCCURRENCE", "빈도"])
    has_detection = any(k in header_str for k in ["검출도", "DETECTION", "검출"])
    has_sod_letters = all(f" {k} " in f" {header_str} " for k in ["S", "O", "D"])
    has_fmea_kw = any(k in header_str for k in ["고장유형", "FAILURE MODE", "FMEA", "PFMEA", "고장 유형"])
    has_ap = any(k in header_str for k in ["AP", "ACTION PRIORITY"])
    has_rpn = "RPN" in header_str

    return {
        "has_severity": has_severity or has_sod_letters,
        "has_occurrence": has_occurrence or has_sod_letters,
        "has_detection": has_detection or has_sod_letters,
        "has_fmea_kw": has_fmea_kw,
        "has_ap": has_ap,
        "has_rpn": has_rpn,
        "is_likely_fmea": has_fmea_kw or ((has_severity or has_sod_letters) and (has_occurrence or has_sod_letters)),
        "ap_or_rpn": "AP" if has_ap else ("RPN" if has_rpn else "미확인"),
    }


def count_data_rows(ws) -> int:
    count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in row):
            count += 1
    return count


def extract_hints_from_name(name: str) -> dict:
    hints = {"공장명_추정": "", "공정유형_추정": "", "고객사_추정": "", "연도_추정": ""}
    for kw in FACTORY_KEYWORDS:
        if kw in name:
            hints["공장명_추정"] = f"{kw}공장"
            break
    for kw in PROCESS_KEYWORDS:
        if kw in name:
            hints["공정유형_추정"] = kw
            break
    for kw in CUSTOMER_KEYWORDS:
        if kw in name:
            hints["고객사_추정"] = kw
            break
    m = re.search(r"20\d{2}", name)
    if m:
        hints["연도_추정"] = m.group()
    return hints


def scan_file(path: Path) -> dict:
    record = {
        "파일명": path.name,
        "경로": str(path.parent),
        "파일크기_KB": round(path.stat().st_size / 1024, 1),
        "수정일": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d"),
        "시트수": 0,
        "데이터행수_추정": 0,
        "FMEA_판별": "X",
        "S_O_D_컬럼": "없음",
        "AP_RPN_방식": "미확인",
        "공장명_추정": "",
        "공정유형_추정": "",
        "고객사_추정": "",
        "연도_추정": "",
        "오류": "",
    }

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        record["시트수"] = len(wb.sheetnames)
        ws = wb.active

        col_info = detect_fmea_columns(ws)
        record["FMEA_판별"] = "O" if col_info["is_likely_fmea"] else "X"
        record["S_O_D_컬럼"] = "있음" if (col_info["has_severity"] and col_info["has_occurrence"]) else "없음"
        record["AP_RPN_방식"] = col_info["ap_or_rpn"]
        record["데이터행수_추정"] = count_data_rows(ws)
        wb.close()
    except Exception as e:
        record["오류"] = str(e)[:120]

    hints = extract_hints_from_name(path.stem)
    record.update(hints)
    return record


def scan_directory(root_dir: str, output_csv: str):
    root = Path(root_dir)
    if not root.exists():
        print(f"오류: 폴더 없음 — {root_dir}")
        return

    files = [f for f in list(root.rglob("*.xlsx")) + list(root.rglob("*.xls"))
             if not f.name.startswith("~$")]

    print(f"[Phase 1] Excel 스캔 시작: {root_dir}")
    print(f"  발견된 파일: {len(files)}개")

    if not files:
        print("  파일 없음. 종료.")
        return

    records = []
    for i, f in enumerate(files, 1):
        print(f"  [{i:3}/{len(files)}] {f.name[:50]}", end="\r", flush=True)
        records.append(scan_file(f))

    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)

    fmea_count = sum(1 for r in records if r["FMEA_판별"] == "O")
    print(f"\n  스캔 완료.")
    print(f"    전체: {len(records)}개  |  FMEA 판별: {fmea_count}개  |  비FMEA: {len(records) - fmea_count}개")
    print(f"\n  CSV 저장: {output_csv}")
    print("  → 다음 단계: python metadata_tagger.py --input <이 CSV>")
    print("  → 담당자 확인: 공장명_추정, 공정유형_추정, 고객사_추정 컬럼 수동 검토 필요")


def main():
    parser = argparse.ArgumentParser(description="Phase 1 — FMEA Excel 파일 스캐너")
    parser.add_argument("--dir", required=True, help="스캔할 폴더 경로 (공유 서버 경로 포함)")
    default_out = f"fmea_목록표_{datetime.now().strftime('%Y%m%d')}.csv"
    parser.add_argument("--output", default=default_out, help=f"출력 CSV 파일명 (기본: {default_out})")
    args = parser.parse_args()

    scan_directory(args.dir, args.output)


if __name__ == "__main__":
    main()
