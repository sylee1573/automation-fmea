#!/usr/bin/env python3
"""
Phase 1: FMEA 품질 필터 + 등급 분류

metadata_tagger.py 출력 CSV를 받아 A/B/C 등급 분류 후 RAG 투입 대상 확정.

실행: python quality_filter.py --input 목록표_tagged.csv [--output 목록표_graded.csv] [--verify]

등급 기준:
  A: FMEA 판별 O, S/O/D 컬럼 있음, 데이터 10행 이상 → RAG 투입
  B: FMEA 판별 O, S/O/D 있음, 데이터 5~9행 → RAG 투입 (가중치 낮게)
  C: FMEA 아님 / S/O/D 없음 / 행 부족 / 폐기·단종 키워드 → RAG 제외

--verify 옵션: 실제 파일 열어 S/O/D 값 채움 비율 추가 검증 (느림, 선택적)
"""

import argparse
import csv
from pathlib import Path

import openpyxl

EXCLUDE_KEYWORDS = ["폐기", "단종", "구버전", "_old", "_bak", "_backup", "_test", "~$"]


def grade_by_metadata(row: dict) -> tuple:
    filename = row.get("파일명", "").lower()
    data_rows = int(row.get("데이터행수_추정", 0) or 0)
    has_sod = row.get("S_O_D_컬럼", "") == "있음"
    is_fmea = row.get("FMEA_판별", "") == "O"
    error = row.get("오류", "")

    for kw in EXCLUDE_KEYWORDS:
        if kw in filename:
            return "C", f"제외 키워드 포함: {kw}"

    if error:
        return "C", f"파일 오류: {error[:50]}"
    if not is_fmea:
        return "C", "FMEA 아님"
    if not has_sod:
        return "C", "S/O/D 컬럼 없음"

    if data_rows >= 10:
        return "A", f"완전한 데이터 ({data_rows}행)"
    elif data_rows >= 5:
        return "B", f"데이터 {data_rows}행 (5~9행)"
    else:
        return "C", f"데이터 부족 ({data_rows}행)"


def verify_file_sod_fill(file_path: str) -> float:
    """실제 파일에서 S/O/D 값 채움 비율 반환 (0.0~1.0)"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        sod_cols = []
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for ci, cell in enumerate(row):
                if isinstance(cell, str) and cell.strip() in ("S", "O", "D"):
                    sod_cols.append(ci)
            if len(sod_cols) >= 2:
                break

        if not sod_cols:
            wb.close()
            return 0.0

        total = filled = 0
        for row in ws.iter_rows(min_row=6, values_only=True):
            if any(c is not None for c in row):
                total += 1
                if sum(1 for ci in sod_cols if ci < len(row) and row[ci] is not None) >= 2:
                    filled += 1

        wb.close()
        return filled / max(total, 1)
    except Exception:
        return 0.0


def main():
    parser = argparse.ArgumentParser(description="Phase 1 — FMEA 품질 필터 + 등급 분류")
    parser.add_argument("--input", required=True, help="metadata_tagger.py 출력 CSV")
    parser.add_argument("--output", default="", help="등급 분류 결과 CSV")
    parser.add_argument("--verify", action="store_true",
                        help="실제 파일 열어 S/O/D 채움 비율 검증 (느림, 권장)")
    args = parser.parse_args()

    if not args.output:
        p = Path(args.input)
        args.output = str(p.parent / f"{p.stem}_graded.csv")

    with open(args.input, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    print(f"[Phase 1] 품질 등급 분류: {len(rows)}개 파일")
    if args.verify:
        print("  --verify 옵션 활성: 실제 파일 S/O/D 채움 비율 추가 검증")

    graded = []
    for i, row in enumerate(rows, 1):
        grade, reason = grade_by_metadata(row)

        if args.verify and grade in ("A", "B"):
            file_path = str(Path(row.get("경로", "")) / row.get("파일명", ""))
            fill_rate = verify_file_sod_fill(file_path)

            if fill_rate < 0.3:
                grade, reason = "C", f"S/O/D 채움 {fill_rate:.0%} (기준 30% 미달)"
            elif fill_rate < 0.7 and grade == "A":
                grade, reason = "B", f"S/O/D 채움 {fill_rate:.0%} (70% 미달 → B 강등)"

        row["품질등급"] = grade
        row["등급_사유"] = reason
        row["RAG_투입"] = "O" if grade in ("A", "B") else "X"
        graded.append(row)
        grade_icon = {"A": "✓", "B": "△", "C": "✗"}.get(grade, "?")
        print(f"  [{i:3}/{len(rows)}] {row.get('파일명','')[:40]:<40} {grade_icon} {grade}: {reason}")

    if graded:
        with open(args.output, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=graded[0].keys())
            writer.writeheader()
            writer.writerows(graded)

    a = sum(1 for r in graded if r["품질등급"] == "A")
    b = sum(1 for r in graded if r["품질등급"] == "B")
    c = sum(1 for r in graded if r["품질등급"] == "C")
    rag = sum(1 for r in graded if r["RAG_투입"] == "O")

    print(f"\n  {'─'*45}")
    print(f"  A (완전):  {a:3}개")
    print(f"  B (보통):  {b:3}개")
    print(f"  C (제외):  {c:3}개")
    print(f"  RAG 투입:  {rag:3}개 (A+B)")
    print(f"  {'─'*45}")
    print(f"\n  결과 저장: {args.output}")
    print("  → 다음 단계: Phase 2 (전처리 + Vector DB 인덱싱)")
    print("  → 담당자: C 등급 사유 확인 후 재분류 필요 항목 직접 수정 가능")


if __name__ == "__main__":
    main()
